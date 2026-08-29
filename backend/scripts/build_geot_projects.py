"""Build the compact GEOT project-portfolio artifact from the GEOT xlsx.

The Global Energy Ownership Tracker (GEOT) xlsx is published by Global Energy
Monitor under CC BY 4.0 but sits behind a download form (reCAPTCHA), so it
cannot be fetched at runtime. This script is run manually against each new
GEOT release (roughly twice a year) and the generated artifact is committed:

    python scripts/build_geot_projects.py \
        ~/Downloads/Global-Energy-Ownership-Tracker-August-2026-V2.xlsx \
        --release "August 2026"

Output: ``opencheck/data/geot_projects.json.gz`` with two data sections.

``entities`` — a per-entity summary of the precomputed ownership closure in
the 9 per-tracker sheets:

* Each sheet row is one (ultimate parent, asset unit, ownership path) with an
  *effective* share (product along the path). Shares for one unit sum well
  over 100% by design (every level of each chain is enumerated, plus minority
  shareholders), so rows must never be summed naively.
* "Projects" are distinct project-level asset IDs (plant location / mine /
  pipeline / steel plant / cement plant), not units.
* A project counts as *controlled* when the parent's effective share —
  deduped paths summed per unit, capped at 100, max across the project's
  units — is ≥ 50%.
* Each (parent, project) pair is assigned its best status by priority
  operating > development > mothballed > retired > cancelled > other, so the
  status counts sum to the total distinct projects.

``entity_status`` (August 2026 release onwards) — a compact per-entity record
of the All Entities sheet's corporate-lifecycle columns, present only for
entities that carry any of them:

* ``status`` — ``dissolved`` or ``amalgamated`` (``Entity Status`` column;
  blank in the sheet means active and yields no record).
* ``merged_into`` / ``merged_into_name`` / ``merged_into_lei`` — the successor
  GEM entity for amalgamated companies. The raw ``Merged Into`` value carries
  a spurious ``.0`` float suffix on 115 of 117 rows (August 2026), which is
  stripped; name and LEI are resolved from the same sheet at build time and
  omitted when the successor ID has no row (33 dangle in August 2026).
* ``urls`` — ``Entity Status Data Source URL``, published as a stringified
  Python list (``"['https://…', 'https://…']"``) and parsed accordingly.
* ``jv: true`` — the ``Joint Venture`` column; the literal ``"Unknown"``
  (4 rows in August 2026) is treated as absent.

Requires openpyxl (tooling-only dependency, not needed at runtime):
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import ast
import datetime as dt
import gzip
import json
from collections import defaultdict
from pathlib import Path

# Sheet name → (parent ID column, project-level asset ID column, unit-level
# asset ID column, status column, short tracker key).
# Column naming is inconsistent across sheets — verified against August 2026
# (which renamed the coal sheet's "Owner GEM Entity ID" to "Parent GEM Entity
# ID" and the cement sheet's "GEM Plant ID" to "GEM plant ID"; lookups are
# case-insensitive via _col so pure case drift no longer breaks the build).
SHEET_SPEC: dict[str, tuple[str, str, str, str, str]] = {
    "Coal Plant Ownership": (
        "Parent GEM Entity ID", "GEM location ID", "GEM unit ID", "Status", "coal_plant",
    ),
    "Gas Plant Ownership": (
        "Parent GEM Entity ID", "GEM location ID", "GEM unit ID", "Status", "gas_plant",
    ),
    "Bioenergy Power Ownership": (
        "Parent GEM Entity ID", "GEM location ID", "GEM unit ID", "Status", "bioenergy",
    ),
    "Coal Mine Ownership": (
        "Parent GEM Entity ID", "GEM Mine ID", "GEM Mine ID", "Status", "coal_mine",
    ),
    "Iron Mine Ownership": (
        "Parent GEM Entity ID", "GEM Asset ID", "GEM Asset ID", "Operating status", "iron_mine",
    ),
    "Gas Pipeline Ownership": (
        "Parent GEM Entity ID", "ProjectID", "ProjectID", "Status", "gas_pipeline",
    ),
    "Oil & NGL Pipeline Ownership": (
        "Parent GEM Entity ID", "ProjectID", "ProjectID", "Status", "oil_ngl_pipeline",
    ),
    "Steel Plant Ownership": (
        "Parent GEM Entity ID", "Steel Plant ID", "Steel Plant ID", "Status", "steel_plant",
    ),
    "Cement and Concrete Ownership": (
        "Parent GEM Entity ID", "GEM plant ID", "GEM plant ID", "Status", "cement",
    ),
}

CONTROL_THRESHOLD = 50.0

# Status → bucket. Anything unmapped lands in "other".
STATUS_BUCKETS: dict[str, str] = {
    "operating": "operating",
    "operating pre-retirement": "operating",
    "construction": "development",
    "in construction": "development",
    "pre-construction": "development",
    "permitted": "development",
    "pre-permit": "development",
    "announced": "development",
    "proposed": "development",
    "in development": "development",
    "mothballed": "mothballed",
    "idle": "mothballed",
    "retired": "retired",
    "retired - inferred 2 y": "retired",
    "retired - inferred 4 y": "retired",
    "closed": "retired",
    "cancelled": "cancelled",
    "cancelled - inferred 2 y": "cancelled",
    "cancelled - inferred 4 y": "cancelled",
    "shelved": "cancelled",
    "shelved - inferred 2 y": "cancelled",
}
BUCKET_PRIORITY = ["operating", "development", "mothballed", "retired", "cancelled", "other"]
LIVE_BUCKETS = {"operating", "development"}


def _bucket(status: str | None) -> str:
    return STATUS_BUCKETS.get((status or "").strip().lower(), "other")


def _col(header: list[str], sheet: str, name: str, *, required: bool = True) -> int | None:
    """Index of a column by case-insensitive name, or a legible error.

    GEM's header casing drifts between releases (August 2026 turned
    "GEM Plant ID" into "GEM plant ID"), so exact matching is too brittle;
    a genuinely missing/renamed column should fail loudly, naming the sheet.
    """
    lowered = [h.lower() for h in header]
    try:
        return lowered.index(name.lower())
    except ValueError:
        if not required:
            return None
        raise SystemExit(
            f"Sheet {sheet!r} has no column matching {name!r} — headers: {header}"
        ) from None


# --- All Entities sheet: corporate-lifecycle columns (August 2026 onwards) ---

_AE_SHEET = "All Entities"
_AE_ID = "Entity ID"
_AE_NAME = "Full Name"
_AE_LEI = "Global Legal Entity Identifier Index"
_AE_JV = "Joint Venture"
_AE_STATUS = "Entity Status"
_AE_MERGED = "Merged Into"
_AE_STATUS_URL = "Entity Status Data Source URL"


def _clean_lei(raw: object) -> str | None:
    """A valid 20-char LEI from the sheet's LEI cell, else None.

    The column also holds "not found", "n/a" and semicolon-delimited multiples;
    take the first valid-looking code, mirroring the runtime adapter.
    """
    for token in str(raw or "").split(";"):
        lei = token.strip().upper()
        if len(lei) == 20 and lei not in ("", "NOT FOUND", "N/A"):
            return lei
    return None


def _parse_status_urls(raw: object) -> list[str]:
    """Parse the Entity Status Data Source URL cell.

    August 2026 publishes stringified Python lists ("['https://a', 'https://b']");
    tolerate a plain URL string too in case the format is fixed upstream.
    """
    text = str(raw or "").strip()
    if not text:
        return []
    if text.startswith("["):
        try:
            parsed = ast.literal_eval(text)
            if isinstance(parsed, (list, tuple)):
                return [str(u).strip() for u in parsed if str(u).strip()]
        except (ValueError, SyntaxError):
            pass
    return [text]


def extract_entity_status(wb: object) -> dict[str, dict]:
    """Per-entity corporate-lifecycle records from the All Entities sheet.

    Returns {gem_entity_id: record} for entities that are status-flagged
    (dissolved/amalgamated) or joint ventures; everything else is omitted.
    Pre-August-2026 workbooks lack the columns — returns {} then.
    """
    try:
        ws = wb[_AE_SHEET]
    except KeyError:
        print(f"No {_AE_SHEET!r} sheet — skipping entity_status")
        return {}

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    id_i = _col(header, _AE_SHEET, _AE_ID)
    name_i = _col(header, _AE_SHEET, _AE_NAME)
    lei_i = _col(header, _AE_SHEET, _AE_LEI, required=False)
    jv_i = _col(header, _AE_SHEET, _AE_JV, required=False)
    status_i = _col(header, _AE_SHEET, _AE_STATUS, required=False)
    merged_i = _col(header, _AE_SHEET, _AE_MERGED, required=False)
    url_i = _col(header, _AE_SHEET, _AE_STATUS_URL, required=False)
    if status_i is None and jv_i is None:
        print("All Entities sheet predates the entity-status columns — skipping")
        return {}

    names: dict[str, str] = {}
    leis: dict[str, str] = {}
    records: dict[str, dict] = {}

    for row in rows:
        eid = str(row[id_i] or "").strip()
        if not eid:
            continue
        names[eid] = str(row[name_i] or "").strip()
        if lei_i is not None:
            lei = _clean_lei(row[lei_i])
            if lei:
                leis[eid] = lei

        rec: dict = {}
        status = str(row[status_i] or "").strip().lower() if status_i is not None else ""
        if status in ("dissolved", "amalgamated"):
            rec["status"] = status
            if merged_i is not None:
                # 115 of 117 August-2026 values carry a float-coercion ".0" suffix.
                merged = str(row[merged_i] or "").strip().removesuffix(".0")
                if merged:
                    rec["merged_into"] = merged
            if url_i is not None:
                urls = _parse_status_urls(row[url_i])
                if urls:
                    rec["urls"] = urls
        # The literal "Unknown" (4 rows in August 2026) is not an assertion.
        if jv_i is not None and str(row[jv_i] or "").strip().lower() == "true":
            rec["jv"] = True
        if rec:
            records[eid] = rec

    # Resolve successor name/LEI now so the runtime needs no extra index.
    # 33 August-2026 successor IDs have no All Entities row — keep the bare ID.
    for rec in records.values():
        merged = rec.get("merged_into")
        if not merged:
            continue
        if names.get(merged):
            rec["merged_into_name"] = names[merged]
        if merged in leis:
            rec["merged_into_lei"] = leis[merged]

    n_status = sum(1 for r in records.values() if "status" in r)
    n_jv = sum(1 for r in records.values() if r.get("jv"))
    print(f"{_AE_SHEET:32s} status-flagged={n_status:4d} joint ventures={n_jv:4d}")
    return records


def build(xlsx_path: Path, release: str) -> dict:
    import openpyxl  # tooling-only dependency

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # (parent, tracker, project) → {"units": {unit: share_sum_or_None},
    #                               "bucket": best status bucket}
    projects: dict[tuple[str, str, str], dict] = {}

    for sheet, (pcol, projcol, unitcol, scol, tracker) in SHEET_SPEC.items():
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        idx = {name: _col(header, sheet, name) for name in (pcol, projcol, unitcol, scol)}
        path_idx = _col(header, sheet, "Ownership Path", required=False)
        share_idx = _col(header, sheet, "Share")

        seen_paths: set[tuple] = set()  # exact-duplicate row guard
        n_rows = 0
        for row in rows:
            parent = (str(row[idx[pcol]] or "")).strip()
            project = (str(row[idx[projcol]] or "")).strip()
            unit = (str(row[idx[unitcol]] or "")).strip() or project
            if not parent or not project:
                continue
            n_rows += 1

            # Dedupe fully identical (parent, unit, path) rows — the coal
            # sheet alone has ~700 exact duplicates.
            path_val = row[path_idx] if path_idx is not None else None
            dedupe_key = (parent, unit, str(path_val or ""), str(row[share_idx] or ""))
            if dedupe_key in seen_paths:
                continue
            seen_paths.add(dedupe_key)

            share_raw = row[share_idx]
            try:
                share: float | None = float(share_raw) if share_raw not in (None, "") else None
            except (TypeError, ValueError):
                share = None

            bucket = _bucket(str(row[idx[scol]] or ""))

            key = (parent, tracker, project)
            rec = projects.get(key)
            if rec is None:
                rec = {"units": {}, "bucket": bucket}
                projects[key] = rec
            else:
                if BUCKET_PRIORITY.index(bucket) < BUCKET_PRIORITY.index(rec["bucket"]):
                    rec["bucket"] = bucket
            # Distinct paths to the same unit add up (e.g. two 30% routes),
            # capped at 100. Unknown shares stay unknown.
            if share is not None:
                prev = rec["units"].get(unit)
                rec["units"][unit] = min(100.0, (prev or 0.0) + share)
            else:
                rec["units"].setdefault(unit, None)
        print(f"{sheet:32s} rows={n_rows:6d}")

    # Aggregate per entity.
    entities: dict[str, dict] = {}
    for (parent, tracker, _project), rec in projects.items():
        ent = entities.setdefault(
            parent,
            {
                "total": [0, 0, 0],  # [live, operating, controlled (live, ≥50%)]
                "statuses": defaultdict(int),
                "trackers": defaultdict(lambda: [0, 0, 0]),
            },
        )
        bucket = rec["bucket"]
        ent["statuses"][bucket] += 1

        shares = [s for s in rec["units"].values() if s is not None]
        controlled = bool(shares) and max(shares) >= CONTROL_THRESHOLD

        if bucket in LIVE_BUCKETS:
            t = ent["trackers"][tracker]
            t[0] += 1
            if bucket == "operating":
                t[1] += 1
            if controlled:
                t[2] += 1
            ent["total"][0] += 1
            if bucket == "operating":
                ent["total"][1] += 1
            if controlled:
                ent["total"][2] += 1

    out_entities = {
        eid: {
            "total": e["total"],
            "statuses": dict(e["statuses"]),
            "trackers": {k: v for k, v in sorted(e["trackers"].items())},
        }
        for eid, e in entities.items()
    }

    entity_status = extract_entity_status(wb)

    return {
        "meta": {
            "release": release,
            "generated": dt.date.today().isoformat(),
            "source": (
                "Global Energy Ownership Tracker, Global Energy Monitor "
                f"({release} release), CC BY 4.0"
            ),
            "control_threshold_pct": CONTROL_THRESHOLD,
            "live_buckets": sorted(LIVE_BUCKETS),
            "entity_count": len(out_entities),
            "entity_status_count": len(entity_status),
        },
        "entities": out_entities,
        "entity_status": entity_status,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="Path to the GEOT release xlsx")
    ap.add_argument("--release", required=True, help='e.g. "May 2026"')
    ap.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent.parent
        / "opencheck" / "data" / "geot_projects.json.gz",
    )
    args = ap.parse_args()

    data = build(args.xlsx, args.release)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(args.out, "wt", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))
    size_kb = args.out.stat().st_size / 1024
    print(
        f"\nWrote {args.out} ({size_kb:.0f} KB, {data['meta']['entity_count']} entities, "
        f"{data['meta']['entity_status_count']} entity-status records)"
    )


if __name__ == "__main__":
    main()
